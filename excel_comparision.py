from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
before_file = "Before.xlsx"
after_file = "After.xlsx"

wb_before = load_workbook(before_file)
wb_after = load_workbook(after_file)
sheet_before = wb_before.active
sheet_after = wb_after.active
for row in range(1, sheet_after.max_row + 1):
    for col in range(1, sheet_after.max_column + 1):
        cell_before = sheet_before.cell(row=row, column=col)
        cell_after = sheet_after.cell(row=row, column=col)

        if cell_before.value != cell_after.value:
            fill = PatternFill(fill_type="solid", fgColor="FDE366")
            sheet_after.cell(row=row, column=col).fill = fill
modified_after_file = "modified_after_file.xlsx"
wb_after.save(modified_after_file)
